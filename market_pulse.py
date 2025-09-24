# market_pulse.py
# Created: 2024-12-19
# Last Edited: 2024-12-19
#
# Purpose:
# Daily Market Pulse Report - orchestrates slopes_benchmark and coil_spring
# to generate executive summary for strategic market overview.

import pandas as pd
import numpy as np
import yaml
from pathlib import Path
from datetime import datetime
import sys
import os
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console

# Add src directory to path for indicators import
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from indicators import compute_features

# Import existing scanners
from ma_slopes_scan import load_yaml, load_symbols_from_cli, parse_args

class MarketPulseGenerator:
    """Generates daily market pulse reports with executive summary."""
    
    def __init__(self, config_path: str = "market_pulse.yaml"):
        """Initialize with market pulse configuration."""
        self.config = self.load_config(config_path)
        self.console = Console()
        self.results = {}
        
    def load_config(self, config_path: str) -> dict:
        """Load market pulse configuration."""
        try:
            with open(config_path, 'r') as file:
                config = yaml.safe_load(file)
            return config
        except FileNotFoundError:
            print(f"❌ Configuration file not found: {config_path}")
            return None
        except yaml.YAMLError as e:
            print(f"❌ Error parsing YAML configuration: {e}")
            return None
    
    def run_slopes_benchmark(self) -> pd.DataFrame:
        """Run slopes benchmark and return results."""
        print("🔍 Running slopes benchmark...")
        
        # Load slopes configuration
        slopes_config = load_yaml("slopes_benchmark.yaml")
        if not slopes_config:
            raise ValueError("Failed to load slopes_benchmark.yaml")
        
        # Get symbols
        symbols = self.get_symbols()
        
        # Run slopes analysis for each timeframe
        all_results = []
        timeframes = self.config.get('exec_summary', {}).get('timeframes', ['1h', '4h', '1d'])
        
        for tf in timeframes:
            print(f"  📊 Processing {tf} timeframe...")
            tf_results = self.analyze_timeframe(tf, symbols, slopes_config)
            tf_results['timeframe'] = tf
            all_results.append(tf_results)
        
        # Combine all timeframes
        combined_results = pd.concat(all_results, ignore_index=True)
        return combined_results
    
    def analyze_timeframe(self, timeframe: str, symbols: List[str], config: dict) -> pd.DataFrame:
        """Analyze slopes for a specific timeframe."""
        # Load OHLCV data
        parquet_path = f"ohlcv_parquet/ohlcv_{timeframe}.parquet"
        if not Path(parquet_path).exists():
            print(f"⚠️  OHLCV data not found for {timeframe}: {parquet_path}")
            return pd.DataFrame()
        
        df = pd.read_parquet(parquet_path)
        
        # Filter symbols
        available_symbols = df['symbol'].unique()
        filtered_symbols = [s for s in symbols if s in available_symbols]
        
        if len(filtered_symbols) == 0:
            print(f"⚠️  No symbols found for {timeframe}")
            return pd.DataFrame()
        
        # Compute features for each symbol
        results = []
        for symbol in filtered_symbols:
            try:
                symbol_data = df[df['symbol'] == symbol].copy()
                if len(symbol_data) < 100:  # Need enough data for slopes
                    continue
                
                # Compute features
                try:
                    features = compute_features(symbol_data, config)
                except Exception as e:
                    print(f"⚠️  Error computing features for {symbol} on {timeframe}: {e}")
                    continue
                
                # Extract slope information
                result = {
                    'symbol': symbol,
                    'sma150_slope_pct': features.get('SMA150_SLOPE_BPS', 0) / 100,
                    'ema21_slope_pct': features.get('EMA21_SLOPE_BPS', 0) / 100,
                    'ema40_slope_pct': features.get('EMA40_SLOPE_BPS', 0) / 100,
                    'sma50_slope_pct': features.get('SMA50_SLOPE_BPS', 0) / 100,
                    'close': symbol_data['close'].iloc[-1],
                    'volume': symbol_data['volume'].iloc[-1]
                }
                results.append(result)
                
            except Exception as e:
                print(f"⚠️  Error processing {symbol} on {timeframe}: {e}")
                continue
        
        return pd.DataFrame(results)
    
    def get_symbols(self) -> List[str]:
        """Get symbols from configuration."""
        universe = self.config.get('universe', {})
        exclude_symbols = universe.get('exclude_symbols', [])
        
        # Load all symbols from parquet
        parquet_path = "ohlcv_parquet/ohlcv_1h.parquet"
        if not Path(parquet_path).exists():
            raise FileNotFoundError(f"OHLCV data not found: {parquet_path}")
        
        df = pd.read_parquet(parquet_path)
        all_symbols = df['symbol'].unique().tolist()
        
        # Filter out excluded symbols
        filtered_symbols = [s for s in all_symbols if s not in exclude_symbols]
        
        print(f"📈 Found {len(filtered_symbols)} symbols (excluded {len(exclude_symbols)} stablecoins)")
        return filtered_symbols
    
    def calculate_trend_summary(self, results: pd.DataFrame) -> Dict:
        """Calculate trend summary statistics."""
        if len(results) == 0:
            return {}
        
        deadzone_pct = self.config.get('exec_summary', {}).get('trend_deadzone_pct', 0.25)
        
        summary = {}
        timeframes = results['timeframe'].unique()
        
        for tf in timeframes:
            tf_data = results[results['timeframe'] == tf]
            if len(tf_data) == 0:
                continue
            
            # Use SMA150 slope as primary trend indicator
            slopes = tf_data['sma150_slope_pct']
            
            # Handle NaN values and ensure numeric data
            slopes = pd.to_numeric(slopes, errors='coerce').dropna()
            
            if len(slopes) == 0:
                continue
            
            bullish = (slopes > deadzone_pct).sum()
            bearish = (slopes < -deadzone_pct).sum()
            flat = ((slopes >= -deadzone_pct) & (slopes <= deadzone_pct)).sum()
            total = len(slopes)
            
            summary[tf] = {
                'total': total,
                'bullish': bullish,
                'bearish': bearish,
                'flat': flat,
                'bullish_pct': (bullish / total * 100) if total > 0 else 0,
                'bearish_pct': (bearish / total * 100) if total > 0 else 0,
                'flat_pct': (flat / total * 100) if total > 0 else 0
            }
        
        return summary
    
    def get_top_performers(self, results: pd.DataFrame, count: int = 20) -> Dict:
        """Get top performers by slope strength."""
        if len(results) == 0:
            return {}
        
        top_performers = {}
        timeframes = results['timeframe'].unique()
        
        for tf in timeframes:
            tf_data = results[results['timeframe'] == tf]
            if len(tf_data) == 0:
                continue
            
            # Sort by SMA150 slope strength (absolute value)
            tf_data = tf_data.copy()
            tf_data['slope_strength'] = pd.to_numeric(tf_data['sma150_slope_pct'], errors='coerce').abs()
            
            # Remove rows with NaN slope strength
            tf_data = tf_data.dropna(subset=['slope_strength'])
            
            if len(tf_data) == 0:
                continue
                
            top_tf = tf_data.nlargest(count, 'slope_strength')
            
            top_performers[tf] = top_tf[['symbol', 'sma150_slope_pct', 'close', 'volume']].to_dict('records')
        
        return top_performers
    
    def generate_executive_summary(self, trend_summary: Dict, top_performers: Dict) -> str:
        """Generate executive summary text."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        summary_lines = [
            f"📊 MARKET PULSE REPORT - {timestamp}",
            "=" * 50,
            "",
            "🎯 MARKET HEALTH SUMMARY:",
            ""
        ]
        
        # Add trend summary for each timeframe
        for tf in ['1d', '4h', '1h']:  # Order by importance
            if tf in trend_summary:
                data = trend_summary[tf]
                summary_lines.extend([
                    f"📈 {tf.upper()} TIMEFRAME:",
                    f"   • Bullish: {data['bullish_pct']:.1f}% ({data['bullish']}/{data['total']})",
                    f"   • Bearish: {data['bearish_pct']:.1f}% ({data['bearish']}/{data['total']})",
                    f"   • Flat: {data['flat_pct']:.1f}% ({data['flat']}/{data['total']})",
                    ""
                ])
        
        # Add top performers
        summary_lines.extend([
            "⚡ TOP PERFORMERS:",
            ""
        ])
        
        for tf in ['1d', '4h', '1h']:
            if tf in top_performers and top_performers[tf]:
                summary_lines.append(f"📊 {tf.upper()} Leaders:")
                for i, perf in enumerate(top_performers[tf][:5], 1):  # Top 5 only
                    slope_pct = perf['sma150_slope_pct']
                    direction = "📈" if slope_pct > 0 else "📉"
                    summary_lines.append(f"   {i}. {direction} {perf['symbol']}: {slope_pct:+.2f}%")
                summary_lines.append("")
        
        return "\n".join(summary_lines)
    
    def create_excel_report(self, results: pd.DataFrame, trend_summary: Dict, top_performers: Dict) -> str:
        """Create Excel report with executive summary worksheet."""
        timestamp = datetime.now().strftime("%Y-%m-%d")
        filename = f"market_pulse_{timestamp}.xlsx"
        filepath = Path(self.config.get('outputs', {}).get('out_dir', 'out')) / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create Executive Summary worksheet
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        
        # Formatting styles
        header_font = Font(size=18, bold=True, color="006600")
        section_font = Font(size=16, bold=True)
        data_font = Font(size=14)
        
        # Add executive summary content
        row = 1
        ws_exec[f'A{row}'] = f"MARKET PULSE REPORT - {timestamp}"
        ws_exec[f'A{row}'].font = header_font
        row += 2
        
        # Market Health Summary
        ws_exec[f'A{row}'] = "MARKET HEALTH SUMMARY"
        ws_exec[f'A{row}'].font = section_font
        row += 1
        
        for tf in ['1d', '4h', '1h']:
            if tf in trend_summary:
                data = trend_summary[tf]
                ws_exec[f'A{row}'] = f"{tf.upper()} Timeframe:"
                ws_exec[f'A{row}'].font = Font(bold=True)
                ws_exec[f'B{row}'] = f"Bullish: {data['bullish_pct']:.1f}% ({data['bullish']}/{data['total']})"
                ws_exec[f'C{row}'] = f"Bearish: {data['bearish_pct']:.1f}% ({data['bearish']}/{data['total']})"
                ws_exec[f'D{row}'] = f"Flat: {data['flat_pct']:.1f}% ({data['flat']}/{data['total']})"
                row += 1
        
        row += 1
        
        # Top Performers
        ws_exec[f'A{row}'] = "TOP PERFORMERS"
        ws_exec[f'A{row}'].font = section_font
        row += 1
        
        for tf in ['1d', '4h', '1h']:
            if tf in top_performers and top_performers[tf]:
                ws_exec[f'A{row}'] = f"{tf.upper()} Leaders:"
                ws_exec[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for i, perf in enumerate(top_performers[tf][:10], 1):
                    slope_pct = perf['sma150_slope_pct']
                    direction = "📈" if slope_pct > 0 else "📉"
                    ws_exec[f'A{row}'] = f"{i}. {direction} {perf['symbol']}"
                    ws_exec[f'B{row}'] = f"{slope_pct:+.2f}%"
                    ws_exec[f'C{row}'] = f"${perf['close']:.4f}"
                    row += 1
                row += 1
        
        # Create detailed data worksheets
        for tf in ['1d', '4h', '1h']:
            tf_data = results[results['timeframe'] == tf]
            if len(tf_data) > 0:
                ws_tf = wb.create_sheet(title=f"Slopes {tf.upper()}")
                
                # Add headers
                headers = ['Symbol', 'SMA150 Slope %', 'EMA21 Slope %', 'EMA40 Slope %', 'SMA50 Slope %', 'Close', 'Volume']
                for col, header in enumerate(headers, 1):
                    cell = ws_tf.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # Add data
                data_rows = tf_data.to_dict('records')
                for row_idx, row_data in enumerate(data_rows, 2):
                    try:
                        ws_tf.cell(row=row_idx, column=1, value=str(row_data['symbol']))
                        ws_tf.cell(row=row_idx, column=2, value=float(row_data['sma150_slope_pct']) if pd.notna(row_data['sma150_slope_pct']) else 0.0)
                        ws_tf.cell(row=row_idx, column=3, value=float(row_data['ema21_slope_pct']) if pd.notna(row_data['ema21_slope_pct']) else 0.0)
                        ws_tf.cell(row=row_idx, column=4, value=float(row_data['ema40_slope_pct']) if pd.notna(row_data['ema40_slope_pct']) else 0.0)
                        ws_tf.cell(row=row_idx, column=5, value=float(row_data['sma50_slope_pct']) if pd.notna(row_data['sma50_slope_pct']) else 0.0)
                        ws_tf.cell(row=row_idx, column=6, value=float(row_data['close']) if pd.notna(row_data['close']) else 0.0)
                        ws_tf.cell(row=row_idx, column=7, value=float(row_data['volume']) if pd.notna(row_data['volume']) else 0.0)
                    except Exception as e:
                        print(f"⚠️  Error writing row {row_idx}: {e}")
                        continue
        
        # Save workbook
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        print(f"📊 Excel report saved: {filepath}")
        return str(filepath)
    
    def generate_text_files(self, top_performers: Dict, exec_summary: str) -> Dict[str, str]:
        """Generate text files for TradingView and mobile use."""
        timestamp = datetime.now().strftime("%Y-%m-%d")
        out_dir = Path(self.config.get('outputs', {}).get('out_dir', 'out'))
        out_dir.mkdir(parents=True, exist_ok=True)
        
        files = {}
        
        # Top 20 symbols file (TradingView format)
        symbols_file = out_dir / f"market_pulse_top20_symbols.txt"
        with open(symbols_file, 'w') as f:
            for tf in ['1d', '4h', '1h']:
                if tf in top_performers and top_performers[tf]:
                    f.write(f"# {tf.upper()} Top Performers\n")
                    for perf in top_performers[tf][:20]:
                        f.write(f"binance:{perf['symbol'].lower()}\n")
                    f.write("\n")
        
        files['symbols'] = str(symbols_file)
        
        # Top 20 detailed file
        detailed_file = out_dir / f"market_pulse_top20_detailed.txt"
        with open(detailed_file, 'w') as f:
            for tf in ['1d', '4h', '1h']:
                if tf in top_performers and top_performers[tf]:
                    f.write(f"=== {tf.upper()} TOP PERFORMERS ===\n")
                    for i, perf in enumerate(top_performers[tf][:20], 1):
                        slope_pct = perf['sma150_slope_pct']
                        direction = "📈" if slope_pct > 0 else "📉"
                        f.write(f"{i:2d}. {direction} {perf['symbol']:12s} {slope_pct:+6.2f}% ${perf['close']:8.4f}\n")
                    f.write("\n")
        
        files['detailed'] = str(detailed_file)
        
        # Executive summary file
        summary_file = out_dir / f"market_pulse_exec_summary.txt"
        with open(summary_file, 'w') as f:
            f.write(exec_summary)
        
        files['summary'] = str(summary_file)
        
        print(f"📄 Text files generated:")
        for file_type, filepath in files.items():
            print(f"   • {file_type}: {filepath}")
        
        return files
    
    def run(self) -> Dict:
        """Run the complete market pulse analysis."""
        print("🚀 Starting Market Pulse Analysis...")
        
        try:
            # Run slopes benchmark
            results = self.run_slopes_benchmark()
            
            if len(results) == 0:
                print("❌ No results generated")
                return {}
            
            # Calculate summaries
            trend_summary = self.calculate_trend_summary(results)
            top_performers = self.get_top_performers(results)
            
            # Generate executive summary
            exec_summary = self.generate_executive_summary(trend_summary, top_performers)
            
            # Create outputs
            excel_file = self.create_excel_report(results, trend_summary, top_performers)
            text_files = self.generate_text_files(top_performers, exec_summary)
            
            print("\n✅ Market Pulse Analysis Complete!")
            print(f"📊 Excel Report: {excel_file}")
            print(f"📄 Text Files: {len(text_files)} files generated")
            
            return {
                'excel_file': excel_file,
                'text_files': text_files,
                'trend_summary': trend_summary,
                'top_performers': top_performers,
                'exec_summary': exec_summary
            }
            
        except Exception as e:
            import traceback
            print(f"❌ Error during market pulse analysis: {e}")
            print(f"📋 Traceback: {traceback.format_exc()}")
            return {}

def main():
    """Main entry point."""
    generator = MarketPulseGenerator()
    results = generator.run()
    
    if results:
        print("\n🎯 Executive Summary Preview:")
        print("-" * 50)
        print(results['exec_summary'])

if __name__ == "__main__":
    main()
