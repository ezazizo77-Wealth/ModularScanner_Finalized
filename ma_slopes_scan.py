# ma_slopes_scan.py
from pathlib import Path
from typing import Dict, List, Optional
import re
import yaml
import pandas as pd
from rich.console import Console
from rich.table import Table
import os
from datetime import datetime
import argparse
from copy import deepcopy
import openpyxl
from openpyxl.styles import PatternFill

# --- import path bootstrap (lets us find src/indicators.py) ---
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
CANDIDATES = [
    HERE,                 # project root if indicators.py is alongside
    HERE / "src",         # common layout: src/indicators.py
    HERE.parent,          # parent (in case you run from a subdir)
    HERE.parent / "src",
]

for p in CANDIDATES:
    if (p / "indicators.py").exists():
        sys.path.insert(0, str(p))
    if (p / "src" / "indicators.py").exists():
        sys.path.insert(0, str(p / "src"))
# --------------------------------------------------------------

# Uses your existing indicators.py (you added EMA21/SMA40/SMA50/SMA150 + *_SLOPE_BPS inside compute_features)
try:
    from indicators import compute_features
except Exception as e:
    print(f"[FATAL] Unable to import indicators.compute_features: {e}", file=sys.stderr)
    sys.exit(1)

# ---------- Config & IO ----------

DEFAULT_TFS = ["1h", "4h", "1d"]                         # if timeframes not specified in yaml
DEFAULT_PARQUET = "ohlcv_parquet/ohlcv_{tf}.parquet"     # standard layout
DEFAULT_DEADZONE_PCT = 0.25                              # |slope| < 0.25% -> flat "FLAT"

# ---------- CLI Argument Parsing ----------

def parse_args():
    p = argparse.ArgumentParser(description="MA Slopes Scanner")
    p.add_argument("yaml", nargs="?", default="slopes_benchmark.yaml",
                   help="Path to YAML config (default: slopes_benchmark.yaml)")
    # per-timeframe lookback overrides
    p.add_argument("--lb-1h", type=int, default=None, help="Lookback bars for 1h slopes (overrides YAML slope_window)")
    p.add_argument("--lb-4h", type=int, default=None, help="Lookback bars for 4h slopes (overrides YAML slope_window)")
    p.add_argument("--lb-1d", type=int, default=None, help="Lookback bars for 1d slopes (overrides YAML slope_window)")
    # symbol subset
    p.add_argument("--symbols", type=str, default="",
                   help='Comma-separated symbols to include (e.g. "BTC/USDT,ETH/USDT")')
    p.add_argument("--symbols-file", type=str, default="",
                   help="Path to txt/csv with a 'symbol' column or one symbol per line")
    # slope filtering
    p.add_argument("--min-slope", type=float, default=None,
                   help="Minimum slope percentage to include (e.g. 1.0 for 1%)")
    p.add_argument("--max-slope", type=float, default=None,
                   help="Maximum slope percentage to include (e.g. 5.0 for 5%)")
    p.add_argument("--top-n", type=int, default=None,
                   help="Show only top N symbols by slope strength")
    return p

def load_symbols_from_cli(args_symbols: str, symbols_file: str) -> list[str]:
    out = []
    if args_symbols.strip():
        out.extend([s.strip() for s in args_symbols.split(",") if s.strip()])
    if symbols_file.strip():
        p = Path(symbols_file)
        if not p.exists():
            raise FileNotFoundError(f"--symbols-file not found: {symbols_file}")
        try:
            if p.suffix.lower() == ".csv":
                df = pd.read_csv(p)
                col = next((c for c in df.columns if c.lower()=="symbol"), df.columns[0])
                out.extend(df[col].dropna().astype(str).str.strip().tolist())
            else:
                out.extend([ln.strip() for ln in p.read_text(encoding="utf-8").splitlines() if ln.strip()])
        except Exception as e:
            raise ValueError(f"Error reading symbols file {symbols_file}: {e}")
    
    # normalize + dedupe
    norm = []
    seen = set()
    for s in out:
        a = s.strip()
        if not a: continue
        if a not in seen:
            seen.add(a)
            norm.append(a)
    return norm

def build_whitelist_variants(symbols: list[str]) -> set[str]:
    """
    Accept both 'BTC/USDT' and 'BTCUSDT' forms by expanding the provided list.
    """
    wl = set()
    for s in symbols:
        s1 = s.strip()
        s2 = s1.replace("/", "")
        wl.add(s1)
        wl.add(s2)
    return wl

def load_yaml(yaml_path: str) -> Dict:
    cfg_path = Path(yaml_path)
    if not cfg_path.exists():
        raise FileNotFoundError(f"Config not found: {yaml_path}")
    with cfg_path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def compile_universe_regex(pattern: str) -> re.Pattern:
    return re.compile(pattern or ".*")

def list_timeframes(cfg: Dict) -> List[str]:
    # Optional: if you later add timeframes in yaml, use that; else defaults
    tfs = cfg.get("timeframes")
    if isinstance(tfs, list) and tfs:
        return [str(x) for x in tfs]
    return DEFAULT_TFS

def parquet_path_for(tf: str, cfg: Dict) -> Path:
    # Allow override: paths: { "1h": "...", "4h": "...", "1d": "..." }
    paths = cfg.get("paths") or {}
    path = paths.get(tf) or DEFAULT_PARQUET.format(tf=tf)
    return Path(path)

# ---------- Slope utilities ----------

def direction_from_pct(pct: float, deadzone_pct: float) -> str:
    if pd.isna(pct):
        return ""
    if pct >= deadzone_pct:
        return "UP"
    if pct <= -deadzone_pct:
        return "DOWN"
    return "FLAT"

SLOPE_COLS = [
    ("EMA21_SLOPE_BPS",  "ma21"),
    ("EMA40_SLOPE_BPS",  "ma40"),
    ("SMA50_SLOPE_BPS",  "ma50"),
    ("SMA150_SLOPE_BPS", "ma150"),
]

def compute_tf_features(tf: str, parquet_path: Path, cfg: dict,
                        regex_pat: re.Pattern, exclude_syms: list[str],
                        symbols_whitelist: Optional[set[str]] = None) -> pd.DataFrame:
    if not parquet_path.exists():
        raise FileNotFoundError(f"Parquet not found for {tf}: {parquet_path}")

    df = pd.read_parquet(parquet_path)  # expects: symbol, ts, open, high, low, close, volume
    # Universe filter
    df = df[df["symbol"].astype(str).apply(lambda s: bool(regex_pat.match(s)))]
    if exclude_syms:
        df = df[~df["symbol"].isin(exclude_syms)]

    if symbols_whitelist:
        # accept both slash and noslash forms
        noslash = df["symbol"].str.replace("/", "", regex=False)
        df = df[df["symbol"].isin(symbols_whitelist) | noslash.isin(symbols_whitelist)]

    df = df.sort_values(["symbol", "ts"])

    # compute indicators; compute_features now takes cfg
    df = compute_features(df, cfg)

    # Add data count per symbol for filtering
    data_counts = df.groupby("symbol").size().reset_index(name=f"{tf}_data_count")

    # last row per symbol
    last = df.groupby("symbol", as_index=False).tail(1).reset_index(drop=True)
    
    # Merge with data counts
    last = last.merge(data_counts, on="symbol", how="left")

    keep = ["symbol"] + [c for c, _ in SLOPE_COLS if c in last.columns] + [f"{tf}_data_count"]
    return last[keep]

def merge_on_symbol(frames: List[pd.DataFrame], tf_alias: str) -> pd.DataFrame:
    # rename slope columns with timeframe prefix to avoid collisions
    renamed = []
    for f in frames:
        cols = {c: f"{tf_alias}_{c}" for c in f.columns if c != "symbol"}
        renamed.append(f.rename(columns=cols))
    out = renamed[0]
    for f in renamed[1:]:
        out = out.merge(f, on="symbol", how="outer")
    return out

# ---------- Main Scan ----------

def main(yaml_path: Optional[str] = None):
    console = Console()
    args = parse_args().parse_args()
    yaml_path = args.yaml if yaml_path is None else yaml_path

    cfg = load_yaml(yaml_path)

    control = cfg.get("control", {}) or {}
    deadzone_pct = float(control.get("deadzone_pct", DEFAULT_DEADZONE_PCT))

    uni = cfg.get("universe", {}) or {}
    regex_str = uni.get("regex", ".*")
    exclude = list(uni.get("exclude_symbols", []) or [])
    regex_pat = compile_universe_regex(regex_str)

    # default slope_window from YAML if no override
    yaml_slope_window = int(cfg.get("slope_window", 100))

    # CLI: per-timeframe lookbacks (fallback to YAML)
    lb_map = {
        "1h": args.lb_1h if args.lb_1h is not None else yaml_slope_window,
        "4h": args.lb_4h if args.lb_4h is not None else yaml_slope_window,
        "1d": args.lb_1d if args.lb_1d is not None else yaml_slope_window,
    }

    # CLI: symbol subset
    subset_list = load_symbols_from_cli(args.symbols, args.symbols_file)
    symbols_whitelist = build_whitelist_variants(subset_list) if subset_list else None

    # Determine timeframes for processing
    tfs = list_timeframes(cfg)

    # Process timeframes
    tf_frames = {}
    for tf in tfs:
        pth = parquet_path_for(tf, cfg)

        # Make a per-timeframe copy of cfg with overridden slope_window
        tf_cfg = deepcopy(cfg)
        tf_cfg["slope_window"] = int(lb_map.get(tf, yaml_slope_window))

        console.print(f"[cyan]Loading {tf} → {pth} (lookback={tf_cfg['slope_window']})[/]")
        tf_df = compute_tf_features(
            tf=tf,
            parquet_path=pth,
            cfg=tf_cfg,
            regex_pat=regex_pat,
            exclude_syms=exclude,
            symbols_whitelist=symbols_whitelist
        )
        tf_frames[tf] = tf_df

    # Merge all timeframes on symbol (outer)
    merged = None
    for tf in tfs:
        tf_df = tf_frames[tf]
        # prefix slope columns with timeframe (e.g., "1h_EMA21_SLOPE_BPS")
        # But don't double-prefix the data_count columns
        rename_dict = {}
        for c in tf_df.columns:
            if c != "symbol":
                if c.endswith("_data_count"):
                    rename_dict[c] = c  # Keep data_count columns as-is
                else:
                    rename_dict[c] = f"{tf}_{c}"
        tf_df = tf_df.rename(columns=rename_dict)
        merged = tf_df if merged is None else merged.merge(tf_df, on="symbol", how="outer")

    # Add direction columns per slope with deadzone
    dir_cols = []
    for tf in tfs:
        for slope_col, alias in SLOPE_COLS:
            col = f"{tf}_{slope_col}"
            if col in merged.columns:
                dcol = f"{tf}_{alias}_dir"
                merged[dcol] = merged[col].apply(lambda x: direction_from_pct(x, deadzone_pct))
                dir_cols.append(dcol)

    # Sort by strongest overall slope (sum of all timeframes)
    def calculate_strength(row):
        total_strength = 0
        for tf in tfs:
            for slope_col, _ in SLOPE_COLS:
                col_name = f"{tf}_{slope_col}"
                if col_name in row and pd.notna(row[col_name]):
                    total_strength += abs(row[col_name])
        return total_strength
    












    merged['strength'] = merged.apply(calculate_strength, axis=1)
    sorted_df = merged.sort_values('strength', ascending=False)
    
    # Apply CLI filtering
    filtered_df = sorted_df.copy()
    
    # Apply slope filtering
    if args.min_slope is not None or args.max_slope is not None:
        slope_mask = pd.Series([True] * len(filtered_df), index=filtered_df.index)
        
        for tf in tfs:
            for slope_col, _ in SLOPE_COLS:
                col_name = f"{tf}_{slope_col}"
                if col_name in filtered_df.columns:
                    if args.min_slope is not None:
                        slope_mask &= (filtered_df[col_name].abs() >= args.min_slope)
                    if args.max_slope is not None:
                        slope_mask &= (filtered_df[col_name].abs() <= args.max_slope)
        
        filtered_df = filtered_df[slope_mask]
        console.print(f"[yellow]🔍 Slope filtering applied: {len(filtered_df)} symbols remaining[/]")
    
    # Apply top-N filtering
    if args.top_n is not None:
        filtered_df = filtered_df.head(args.top_n)
        console.print(f"[yellow]🔍 Top-N filtering applied: showing top {args.top_n} symbols[/]")

    # Summary before CSV output
    console.print(f"\n[bold cyan]📊 Processing Summary[/]")
    console.print(f"[green]📁 Config file:[/] {yaml_path}")
    console.print(f"[green]📊 Default slope window:[/] {yaml_slope_window}")
    
    console.print(f"\n[bold yellow]⏱️  Per-Timeframe Lookbacks Used:[/]")
    for tf in tfs:
        lb_value = lb_map.get(tf, yaml_slope_window)
        override_indicator = "🔧 CLI override" if (tf == "1h" and args.lb_1h is not None) or \
                                           (tf == "4h" and args.lb_4h is not None) or \
                                           (tf == "1d" and args.lb_1d is not None) else "📋 YAML default"
        console.print(f"  [cyan]{tf:>3}:[/] {lb_value:>3} bars {override_indicator}")
    
    if symbols_whitelist:
        console.print(f"\n[bold yellow]🎯 Symbol Filtering Applied:[/]")
        console.print(f"  [green]Whitelist:[/] {len(symbols_whitelist)} variants from {len(subset_list)} symbols")
        console.print(f"  [green]Sample:[/] {', '.join(list(symbols_whitelist)[:5])}{'...' if len(symbols_whitelist) > 5 else ''}")
    else:
        console.print(f"\n[bold yellow]🎯 Symbol Filtering:[/] [dim]All symbols processed (no whitelist)[/]")
    
    console.print(f"\n[bold yellow]📈 Deadzone threshold:[/] {deadzone_pct}%")
    console.print(f"[bold yellow]🔄 Timeframes processed:[/] {', '.join(tfs)}")
    console.print(f"[bold yellow]📋 Symbols analyzed:[/] {len(merged)}")
    console.print(f"[dim]────────────────────────────────────────[/]\n")

    # Save Excel file with multiple sheets for better analysis
    os.makedirs("out", exist_ok=True)
    
    # Generate timestamped filename
    timestamp = datetime.utcnow().strftime("%Y-%m-%d___%H-%M")
    out_file = cfg.get("output_csv", f"out/slopes_summary_{timestamp}.xlsx")
    
    # Ensure it goes to out folder if not already specified
    if not out_file.startswith("out/"):
        out_file = f"out/{out_file}"
    
    # Try to create Excel file, fallback to CSV if openpyxl not available
    try:
        # Create Excel writer
        with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        
            # Sheet 1: Main Data (properly formatted slopes + directions)
            formatted_df = filtered_df.copy()
            
            # Convert slope columns to proper percentage format (keep as numbers for Excel)
            for tf in tfs:
                for slope_col, _ in SLOPE_COLS:
                    col_name = f"{tf}_{slope_col}"
                    if col_name in formatted_df.columns:
                        # Convert from percentage to decimal for Excel percentage formatting
                        formatted_df[col_name] = formatted_df[col_name] / 100.0
            
            formatted_df.to_excel(writer, sheet_name='Main_Data', index=False)
            
            # Apply percentage formatting to slope columns
            workbook = writer.book
            worksheet = writer.sheets['Main_Data']
            
            # Color code timeframe headers (row 1)
            timeframe_colors = {
                '1h': 'FFE4B5',  # Light orange
                '4h': 'E6E6FA',  # Light purple
                '1d': 'B0E0E6'   # Light blue
            }
            
            # Find slope columns and apply percentage formatting + header coloring
            for col_idx, col_name in enumerate(formatted_df.columns):
                if '_SLOPE_BPS' in col_name or '_dir' in col_name:
                    # Apply percentage format with 2 decimal places for slope columns
                    if '_SLOPE_BPS' in col_name:
                        for row_idx in range(1, len(formatted_df) + 1):
                            cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                            if cell.value is not None:
                                cell.number_format = '0.00%'
                    
                    # Color code header based on timeframe
                    header_cell = worksheet.cell(row=1, column=col_idx + 1)
                    for tf, color in timeframe_colors.items():
                        if col_name.startswith(f"{tf}_"):
                            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            break
            
            # Add color coding for direction columns
            for col_idx, col_name in enumerate(formatted_df.columns):
                if '_dir' in col_name:
                    for row_idx in range(1, len(formatted_df) + 1):
                        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if cell.value == 'UP':
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif cell.value == 'DOWN':
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                        elif cell.value == 'FLAT':
                            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
            # Sheet 2: Top Performers Analysis
            # Filter out symbols with less than 150 1d data points
            if '1d_data_count' in filtered_df.columns:
                top_20_candidates = filtered_df[filtered_df['1d_data_count'] >= 150].head(20).copy()
                console.print(f"[yellow]🔍 Top Performers filter: {len(top_20_candidates)} symbols with ≥150 1d data points[/]")
            else:
                top_20_candidates = filtered_df.head(20).copy()
                console.print(f"[yellow]⚠️  Warning: 1d data count not available, using top 20 without filter[/]")
            
            analysis_data = []
            
            for _, row in top_20_candidates.iterrows():
                analysis_row = {'symbol': row['symbol'], 'strength': row['strength']}
                
                # Add 1d data count if available
                if '1d_data_count' in row:
                    analysis_row['1d_data_count'] = row['1d_data_count']
                
                # Add slope values and directions for each timeframe
                for tf in tfs:
                    for slope_col, alias in SLOPE_COLS:
                        slope_col_name = f"{tf}_{slope_col}"
                        dir_col_name = f"{tf}_{alias}_dir"
                        
                        if slope_col_name in row:
                            # Convert percentage to decimal for proper Excel formatting
                            analysis_row[f"{tf}_{alias}_slope"] = row[slope_col_name] / 100.0
                            analysis_row[f"{tf}_{alias}_direction"] = row.get(dir_col_name, "")
                
                analysis_data.append(analysis_row)
            
            analysis_df = pd.DataFrame(analysis_data)
            analysis_df.to_excel(writer, sheet_name='Top_Performers', index=False)
            
            # Apply formatting to Top Performers sheet
            top_performers_worksheet = writer.sheets['Top_Performers']
            
            # Apply percentage formatting to slope columns
            for col_idx, col_name in enumerate(analysis_df.columns):
                if '_slope' in col_name:
                    for row_idx in range(1, len(analysis_df) + 1):
                        cell = top_performers_worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if cell.value is not None:
                            cell.number_format = '0.00%'
            
            # Add color coding for direction columns
            for col_idx, col_name in enumerate(analysis_df.columns):
                if '_direction' in col_name:
                    for row_idx in range(1, len(analysis_df) + 1):
                        cell = top_performers_worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if cell.value == 'UP':
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif cell.value == 'DOWN':
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                        elif cell.value == 'FLAT':
                            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            # Sheet 3: Trend Summary
            summary_data = []
            
            # Per-timeframe trend counts
            for tf in tfs:
                up_count = 0
                down_count = 0
                flat_count = 0
                
                for _, alias in SLOPE_COLS:
                    dir_col = f"{tf}_{alias}_dir"
                    if dir_col in filtered_df.columns:
                        up_count += (filtered_df[dir_col] == "UP").sum()
                        down_count += (filtered_df[dir_col] == "DOWN").sum()
                        flat_count += (filtered_df[dir_col] == "FLAT").sum()
                
                summary_data.append({
                    'timeframe': tf,
                    'up_trends': up_count,
                    'down_trends': down_count,
                    'flat_trends': flat_count,
                    'total_signals': up_count + down_count + flat_count,
                    'up_percentage': f"{(up_count / (up_count + down_count + flat_count) * 100):.1f}%" if (up_count + down_count + flat_count) > 0 else "0%"
                })
            
            # Overall summary
            total_signals = len(filtered_df) * len(tfs) * len(SLOPE_COLS)
            total_up = sum(1 for _, row in filtered_df.iterrows() 
                           for tf in tfs for _, alias in SLOPE_COLS
                           if row.get(f"{tf}_{alias}_dir") == "UP")
            total_down = sum(1 for _, row in filtered_df.iterrows() 
                             for tf in tfs for _, alias in SLOPE_COLS
                             if row.get(f"{tf}_{alias}_dir") == "DOWN")
            total_flat = sum(1 for _, row in filtered_df.iterrows() 
                             for tf in tfs for _, alias in SLOPE_COLS
                             if row.get(f"{tf}_{alias}_dir") == "FLAT")
            
            summary_data.append({
                'timeframe': 'OVERALL',
                'up_trends': total_up,
                'down_trends': total_down,
                'flat_trends': total_flat,
                'total_signals': total_signals,
                'up_percentage': f"{(total_up / total_signals * 100):.1f}%" if total_signals > 0 else "0%"
            })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Trend_Summary', index=False)
            
            # Sheet 4: Configuration Used
            config_data = [
                {'parameter': 'config_file', 'value': yaml_path},
                {'parameter': 'default_slope_window', 'value': yaml_slope_window},
                {'parameter': 'deadzone_threshold', 'value': f"{deadzone_pct}%"},
                {'parameter': 'timeframes', 'value': ', '.join(tfs)},
                {'parameter': 'symbols_analyzed', 'value': len(filtered_df)},
                {'parameter': 'min_slope_filter', 'value': args.min_slope if args.min_slope else 'None'},
                {'parameter': 'max_slope_filter', 'value': args.max_slope if args.max_slope else 'None'},
                {'parameter': 'top_n_filter', 'value': args.top_n if args.top_n else 'None'},
            ]
            
            # Add per-timeframe lookback values
            for tf in tfs:
                lb_value = lb_map.get(tf, yaml_slope_window)
                config_data.append({'parameter': f'{tf}_lookback', 'value': lb_value})
            
            config_df = pd.DataFrame(config_data)
            config_df.to_excel(writer, sheet_name='Configuration', index=False)
        
        console.print(f"[green]Saved Excel file →[/] {out_file}")
        console.print(f"[dim]Sheets: Main_Data, Top_Performers, Trend_Summary, Configuration[/]")
        
        # Create .txt file with Top Performers symbols for Binance import
        txt_file = out_file.replace('.xlsx', '_top_performers.txt')
        with open(txt_file, 'w') as f:
            for symbol in top_20_candidates['symbol']:
                # Convert symbol to lowercase and ensure proper format for Binance
                binance_symbol = f"binance:{symbol.lower()}"
                f.write(f"{binance_symbol}\n")
        
        console.print(f"[green]Saved Binance import file →[/] {txt_file}")
        console.print(f"[dim]Contains {len(top_20_candidates)} symbols with ≥150 1d data points[/]")
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        console.print(f"[yellow]⚠️  openpyxl not available, saving as CSV instead[/]")
        
        # Save main data as CSV
        csv_file = out_file.replace('.xlsx', '.csv')
        formatted_df = filtered_df.copy()
        for tf in tfs:
            for slope_col, _ in SLOPE_COLS:
                col_name = f"{tf}_{slope_col}"
                if col_name in formatted_df.columns:
                    # Keep as percentage for CSV (since CSV doesn't have formatting)
                    formatted_df[col_name] = formatted_df[col_name].round(2)
        
        formatted_df.to_csv(csv_file, index=False)
        
        # Save summary as separate CSV
        summary_file = csv_file.replace('.csv', '_summary.csv')
        summary_data = []
        
        for tf in tfs:
            up_count = 0
            down_count = 0
            flat_count = 0
            
            for _, alias in SLOPE_COLS:
                dir_col = f"{tf}_{alias}_dir"
                if dir_col in filtered_df.columns:
                    up_count += (filtered_df[dir_col] == "UP").sum()
                    down_count += (filtered_df[dir_col] == "DOWN").sum()
                    flat_count += (filtered_df[dir_col] == "FLAT").sum()
            
            summary_data.append({
                'timeframe': tf,
                'up_trends': up_count,
                'down_trends': down_count,
                'flat_trends': flat_count,
                'total_signals': up_count + down_count + flat_count,
                'up_percentage': f"{(up_count / (up_count + down_count + flat_count) * 100):.1f}%" if (up_count + down_count + flat_count) > 0 else "0%"
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_csv(summary_file, index=False)
        
        console.print(f"[green]Saved CSV files →[/] {csv_file}")
        console.print(f"[green]Summary file →[/] {summary_file}")
        console.print(f"[dim]💡 Install openpyxl for Excel output: pip install openpyxl[/]")
        
        # Create .txt file with Top Performers symbols for Binance import (CSV fallback)
        # Filter out symbols with less than 150 1d data points
        if '1d_data_count' in filtered_df.columns:
            top_20_candidates_csv = filtered_df[filtered_df['1d_data_count'] >= 150].head(20).copy()
        else:
            top_20_candidates_csv = filtered_df.head(20).copy()
        
        txt_file = csv_file.replace('.csv', '_top_performers.txt')
        with open(txt_file, 'w') as f:
            for symbol in top_20_candidates_csv['symbol']:
                # Convert symbol to lowercase and ensure proper format for Binance
                binance_symbol = f"binance:{symbol.lower()}"
                f.write(f"{binance_symbol}\n")
        
        console.print(f"[green]Saved Binance import file →[/] {txt_file}")
        console.print(f"[dim]Contains {len(top_20_candidates_csv)} symbols with ≥150 1d data points[/]")

if __name__ == "__main__":
    main()